import os
import glob
import time
from datetime import datetime
from openpyxl import Workbook
import serial
import os.path
import smbus
import ms5837
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment 
os.system('modprobe w1-gpio')
os.system('modprobe w1-therm')
 
base_dir = '/sys/bus/w1/devices/'
device_folder = glob.glob(base_dir + '28*')[0]
device_file = device_folder + '/w1_slave'

curr_time=datetime.now()
curr_timestamp=datetime.timestamp(curr_time)
filename=input('Give File Name(Location): ')
file_name = filename + ".xlsx"
workbook=Workbook()
sheet=workbook.active
sheet.merge_cells('A1:D1')
cell=sheet.cell(row=1, column=1)
cell=sheet["A1"]
cell.font = Font(size=10,
		bold=True,
		italic=True,
		underline='single',
		color='008000')
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.value=filename

sheet.merge_cells('A2:D2')
cell=sheet.cell(row=2, column=1)
cell.font = Font(size=9,
		bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center")
today=datetime.now()
cell.value=today

cell=sheet["A3"]
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.value="Timestamp"
cell=sheet["B3"]
cell.alignment=Alignment(horizontal="center", vertical="center")
cell.value="Temperature (C)"
cell=sheet["C3"]
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.value="EC (V)"
cell=sheet["D3"]
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.value="Pressure (MSI)"

i2c_ch = 1
i2c_address = 0x76
sensor=ms5837.MS5837_30BA()

if not sensor.init():
	print("Sensor not initialized")
	exit(1)
x=0


def read_temp_raw():
    f = open(device_file, 'r')
    lines = f.readlines()
    f.close()
    return lines
 
def read_temp():
    lines = read_temp_raw()
    while lines[0].strip()[-3:] != 'YES':
        time.sleep(0.2)
        lines = read_temp_raw()
    equals_pos = lines[1].find('t=')
    if equals_pos != -1:
        temp_string = lines[1][equals_pos+2:]
        temp_c = float(temp_string) / 1000.0
        temp_f = temp_c * 9.0 / 5.0 + 32.0
        return temp_c

if __name__ == '__main__':
	ser=serial.Serial('/dev/ttyUSB0', 9600, timeout=1)
	ser.reset_input_buffer()
time.sleep(2)	
while True:
	print(read_temp())	
	cell_temp  = read_temp()
	current_time = datetime.now()
	if __name__ == '__main__':
		ser=serial.Serial('/dev/ttyUSB0', 9600, timeout=1)
		ser.reset_input_buffer()
		cell_EC=ser.readline().decode('utf-8').rstrip()
		cell_temp=read_temp()
		if sensor.read():
			cell_pressure=sensor.pressure()
		current_time=datetime.now()
		print(cell_temp)
		print(cell_EC)
		print(cell_pressure)
		data=(current_time, cell_temp, cell_EC, cell_pressure)
		sheet.append(data)
		time.sleep(2)
		workbook.save(filename=filename)
workbook.save(filename=file_name)
